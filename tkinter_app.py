#!/usr/bin/python
# -*- coding: utf-8 -*-
# encoding=utf8

import sys
reload(sys)
sys.setdefaultencoding('utf8')

import requests
from bs4 import BeautifulSoup, SoupStrainer
import Tkinter
from lxml import html
import openpyxl
from time import sleep


def center(toplevel):
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth()
    h = toplevel.winfo_screenheight()
    size = tuple(int(_) for _ in toplevel.geometry().split('+')[0].split('x'))
    x = w/2 - size[0]/2
    y = h/2 - size[1]/2
    toplevel.geometry("%dx%d+%d+%d" % (size + (x, y)))


class simpleapp_tk(Tkinter.Tk):
    def __init__(self,parent):
        Tkinter.Tk.__init__(self,parent)
        self.parent = parent
        self.initialize()

    def initialize(self):
        self.grid()

        self.entryVariable = Tkinter.StringVar()
        self.entry = Tkinter.Entry(self,textvariable=self.entryVariable)
        self.entry.grid(column=0,row=0,sticky='EW', padx=10, pady=10)
        self.entry.bind("<Return>", self.OnPressEnter)

        self.entryVariable2 = Tkinter.StringVar()
        self.entry2 = Tkinter.Entry(self,textvariable=self.entryVariable2)
        self.entry2.grid(column=0,row=1,sticky='EW', padx=10, pady=0)
        self.entry2.bind("<Return>", self.OnPressEnter)

        self.entryVariable.set(u"Enter Url here.")
        self.entryVariable2.set(u"Enter File name here.")




        button = Tkinter.Button(self,text=u"Scrape !",
                                command=self.OnButtonClick)
        button.grid(column=1,row=0, padx=10, pady=10)



        self.labelVariable = Tkinter.StringVar()
        label = Tkinter.Label(self,textvariable=self.labelVariable,
                              anchor="w",fg="white",bg="blue", padx=10, pady=10)
        label.grid(column=0,row=2,columnspan=2,sticky='EW', padx=10, pady=10)
        self.labelVariable.set(u"Hello Qman55!")


        self.grid_columnconfigure(0,weight=1)
        self.resizable(True,False)
        self.update()
        self.geometry(self.geometry())       
        self.entry.focus_set()
        self.entry.selection_range(0, Tkinter.END)



    def OnButtonClick(self):
        url = self.entryVariable.get() 
        bookname = self.entryVariable2.get()
        self.labelVariable.set(u"Please wait while the bot scrapes ...")
        this(url, bookname)
        self.labelVariable.set(u"Finished!")
        sleep(10)
        exit()

        self.entry.focus_set()
        self.entry.selection_range(0, Tkinter.END)

    def OnPressEnter(self,event):
        self.labelVariable.set( self.entryVariable.get()+" (You pressed ENTER)" )
        self.entry.focus_set()
        self.entry.selection_range(0, Tkinter.END)



def gh(url):
   
    all_products = []
    prod_pages = []
    
    prod_pages.append(url)

    for page_num in range(1):
      
      try:
        # go get a url
        print(url)
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        #pprint(soup)
        
        #"product-wrapper tile-view"
        products = soup.find_all('a', {"class": "product-link"})
        #print(products)
        
        # GET INDIVIDUAL href from product div tag
        counter = 0
        for product in products:
          if counter < 4 :
            all_products.append(product['href'])
          counter += 1


        # GET INDIVIDUAL href from product div tag
        # for product in products:
        #     all_products.append(product['href'])


        # find the next-page button href
        elem = soup.find("a", {"class": "next"})
        # it becomes the new url
        url = (elem['href'])

      except Exception as e:
        #print(all_products)
        #raise e
        break


    print(str(len(all_products))+ " Total Products of this search...")
    print
    print("'The scraper is now getting info of individual searched products and saving them to your workbook'")
    print

    return all_products


def this(url, xl):
    
    print
    print("'This scraper mines custom search product info from https://www.overstock.com'")
    print("'Please use a file name that doesnot contain spaces, no special characters and please be case sensitive'")
    print

    if not ".xlsx" in xl:
      xl = str(xl) + '.xlsx'
    else:
      pass

    book = openpyxl.Workbook()
    active_sheet = book.active
    active_sheet.title = "overstock"

    rowNum = 1
    heads = ['Title', 'Price1', 'Price2', 'Shipping', 'Return', 'Category', 'Reviews', 'Review Average', 'Low Quantity',
              'OOS', 'New Arrival', 'Clearance', 'Top Seller', 'Exclusive', 'Special', 'Weekly Deals', 'Variations', 'Rewards',
              'Link']

    for colNum in range(1, 20):
      val = heads[colNum-1]
      active_sheet.cell(row=rowNum, column=colNum).value = val


    # scrape them, product by product
    for url in gh(url):
        print(url)
        print('-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=')
        rowNum = rowNum+1
        active_sheet.cell(row=rowNum, column=19).value = str(url)

        try:
          response = requests.get(url)
          soup = BeautifulSoup(response.content, 'html.parser')
        except Exception as e:
          sleep(10)
          response = requests.get(url)
          soup = BeautifulSoup(response.content, 'html.parser')


        try:
            # prod title
            elem_value = (soup.find('h1')).text.strip()
            active_sheet.cell(row=rowNum, column=1).value = str(elem_value)
            elem_value = " ".join(elem_value.split())
            print(elem_value, 'title')
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # price1
            elem_value = soup.find('span', {"class": "monetary-price-value"})['content']
            active_sheet.cell(row=rowNum, column=2).value = str(elem_value)
            print(elem_value, 'price')
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # price2
            elem_value = soup.find('span', {"class": 'reference-price'}).text.strip()
            active_sheet.cell(row=rowNum, column=3).value = str(elem_value)
            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # shipping
            elems = soup.find('div', {"class": 'shipping-returns'})
            h5s = elems.find_all('h5')
            ps = elems.find_all('p')

            for elem in h5s:
              if 'Shipping:' == str(elem.text):
                index = h5s.index(elem)
                elem_value = elem.find_next('p').text.strip()
                print(elem_value, 'shipping')
                active_sheet.cell(row=rowNum, column=4).value = str(elem_value)
              elif 'Standard Return Policy:' in str(elem):
                elem_value = elem.text
                print(elem_value, 'return shipping')
                active_sheet.cell(row=rowNum, column=5).value = str(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # category
            elems = soup.find('ul', {"class": 'breadcrumbs'})
            elems = elems.find_all('li')
            elem_value = elems[2].text.strip()
            print(elem_value)
            active_sheet.cell(row=rowNum, column=6).value = str(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # reviews
            elem_value = soup.find('span', {"class": "count"}).text.strip()
            active_sheet.cell(row=rowNum, column=7).value = str(elem_value)
            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # average review
            elem_value = soup.find('div', {"class": "overall-rating"}).text.strip()
            active_sheet.cell(row=rowNum, column=8).value = str(elem_value)
            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # low quantity verified
            elem_value = soup.find('div', {"class": "sellout-risk"}).text.strip()
            active_sheet.cell(row=rowNum, column=9).value = str(elem_value)
            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # out of stock  -- verified
            elem_value = soup.find('div', {"class": "out-of-stock-label"}).text.strip()
            active_sheet.cell(row=rowNum, column=10).value = str(elem_value)
            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # new arrival -- ver
            elem_values = soup.find_all('div', {"class": "message"})
            
            for elem in elem_values:
                elem_value = elem.text.strip()
                if 'New Arrival' in elem_value:
                  elem_value = 'New Arrival'
                  print(elem_value)
                  active_sheet.cell(row=rowNum, column=11).value = str(elem_value)

                elif 'Flash Deal' in elem_value:
                  elem_value = 'Flash Deal'
                  print(elem_value)


                elif 'Top Seller' in elem_value:
                  elem_value = 'Top Seller'
                  print(elem_value)
                  active_sheet.cell(row=rowNum, column=13).value = str(elem_value)

                elif 'Clearance' in elem_value:
                  elem_value = 'Clearance'
                  print(elem_value)
                  active_sheet.cell(row=rowNum, column=12).value = str(elem_value)

                else:
                  elem_value = ''

            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # exclusive - ver
            elem_value = soup.find('div', {"class": "clickable-icon"}).text.strip()
            if "exclusive" in str(elem_value).lower():
              elem_value = 'Exclusive'
            else:
              elem_value = ''
            active_sheet.cell(row=rowNum, column=14).value = str(elem_value)
            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass



        try:
            # special 
            elem_value = soup.find('div', {"class": "clickable-icon"}).text.strip()
            if "special" in str(elem_value).lower():
              elem_value = 'Special'
            else:
              elem_value = ''
            active_sheet.cell(row=rowNum, column=15).value = str(elem_value)
            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # weekly deals -
            elem_value = soup.find('span', {"class": "clickable-icon"})
            if "weekly" in str(elem_value).lower():
              elem_value = elem_value['title']
            else:
              elem_value = ''
            active_sheet.cell(row=rowNum, column=16).value = str(elem_value)
            print(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # options -- variations
            elem_values = soup.find('div', {"id": "optbreakout"})
            titles = elem_values.find_all('h4')
            titles = [h4.text.strip() for h4 in titles]
            elem_value = " & ".join(titles)
            print(elem_value)
            active_sheet.cell(row=rowNum, column=17).value = str(elem_value)
        except Exception as e:
          print(e) #reason.
          pass


        try:
            # rewards - ver     //*[@id="clubo-container"]/div[3]/p[2]  
            elem_value = soup.find('p', {"class": "co-me-rewards"}).text.strip()
            active_sheet.cell(row=rowNum, column=18).value = str(elem_value)
            print(elem_value)
        except Exception as e:
            print(e) #reason.
            pass

        print("Product " +str(rowNum)+ "'-Saved!'")
        print
        print
        print

    book.save(xl)     





if __name__ == "__main__":

    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.1 Safari/537.36',
         'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
         'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
         'Accept-Encoding': 'none',
         'Accept-Language': 'en-US,en;q=0.8',
         'Connection': 'keep-alive'}

    app = simpleapp_tk(None)
    app.title('OVERSTOCK.COM SCRAPER')
    app.geometry('{}x{}'.format(750, 129))

    center(app)

    app.mainloop()




#!/usr/bin/python
# -*- coding: iso-8859-1 -*-